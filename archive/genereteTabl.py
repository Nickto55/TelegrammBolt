# -*- coding: utf-8 -*-
import pandas as pd
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


import config
from user_manager import get_user_display_name

class MainGen:
    def __init__(self):
        self.DATA_FILE = config.DATA_FILE

        self.data = config.load_data(self.DATA_FILE)

    def main(self):
        result = []
        
        result.append(["Деталь","РЦ", "Тип проблемы", "Описание", "Дата", "Время", "Пользователь"])

        if pd.isna(self.data):
            return [["Нет данных для экспорта"]]

        for data_key in self.data.keys():
            for data_chs in self.data[data_key]:
                result_row = ["" for i in range(7)]  # Исправлено на 7 колонок
                for key_value, value in dict(data_chs).items():
                    if key_value == "dse": result_row[0] = value
                    if key_value == "rc": result_row[1] = value
                    if key_value == "problem_type": result_row[2] = value
                    if key_value == "description": result_row[3] = value
                    if key_value == "datetime":
                        result_row[4] = str(value)[:10]
                        result_row[5] = str(value)[11:19]
                    if key_value == "user_id": 
                        # Используем функцию для получения отображаемого имени
                        result_row[6] = get_user_display_name(value)
                result.append(result_row)
        return result


def auto_fit_columns(sheet):
    for column_cells in sheet.columns:
        koo = 0
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = (max_length + koo)  # Немного увеличим для красоты
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


class ExcelWriter:
    def __init__(self, file_path, min_prog: str = None):
        """
        Инициализация класса для работы с Excel файлом

        :param file_path: Полный путь к Excel файлу
        """
        self.file_path = file_path
        self.min_prog = min_prog

        self.fill_color1 = PatternFill(start_color="6f747c", fill_type="solid")
        self.fill_color2 = PatternFill(start_color="aaadb2", fill_type="solid")
        self.fill_color3 = PatternFill(start_color="d9d9d9", fill_type="solid")
        self.fill_color4 = PatternFill(start_color="808080", fill_type="solid")

        self.alfavit = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']


    def write_to_sheet(self, data, sheet_name, start_row=1, start_col=1):
        """
        Запись данных в указанный лист Excel файла

        :param data: Данные для записи (список списков или DataFrame)
        :param sheet_name: Имя листа
        :param start_row: Начальная строка (по умолчанию 1)
        :param start_col: Начальный столбец (по умолчанию 1)
        """
        try:
            # Если файл существует, загружаем его
            if os.path.exists(self.file_path):
                # Загружаем существующую книгу
                book = load_workbook(self.file_path)

                # Если лист существует, удаляем его
                if sheet_name in book.sheetnames:
                    del book[sheet_name]

                # Создаем новый лист
                sheet = book.create_sheet(sheet_name)

                # Записываем данные
                if isinstance(data, pd.DataFrame):
                    # Если данные в формате DataFrame
                    for r_idx, row in enumerate(data.values, start=start_row):
                        for c_idx, value in enumerate(row, start=start_col):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                else:
                    # Если данные в формате списка списков
                    for r_idx, row in enumerate(data, start=start_row):
                        for c_idx, value in enumerate(row, start=start_col):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                auto_fit_columns(sheet)
                # Сохраняем изменения

                book.save(self.file_path)

            else:
                if isinstance(data, pd.DataFrame):
                    with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                        data.to_excel(writer, sheet_name=sheet_name, startrow=start_row - 1, startcol=start_col - 1,
                                      index=False)
                else:
                    df = pd.DataFrame(data)
                    with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row - 1, startcol=start_col - 1,
                                    index=False)

            print(f"Данные успешно записаны в лист '{sheet_name}' файла {self.file_path}")

        except Exception as e:
            print(f"Ошибка при записи в Excel: {e}")

    def append_to_sheet(self, data, sheet_name):
        """
        Добавление данных в конец существующего листа

        :param data: Данные для добавления
        :param sheet_name: Имя листа
        """
        try:
            if os.path.exists(self.file_path):
                book = load_workbook(self.file_path)

                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    # Определяем следующую строку
                    start_row = sheet.max_row + 1
                else:
                    sheet = book.create_sheet(sheet_name)
                    start_row = 1

                # Записываем данные
                if isinstance(data, pd.DataFrame):
                    for r_idx, row in enumerate(data.values, start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                else:
                    for r_idx, row in enumerate(data, start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                book.save(self.file_path)
                print(f"Данные успешно добавлены в лист '{sheet_name}' файла {self.file_path}")
            else:
                # Если файла нет, просто создаем его
                self.write_to_sheet(data, sheet_name)

        except Exception as e:
            print(f"Ошибка при добавлении данных в Excel: {e}")


if __name__=="__main__":
    run = MainGen()
    exl = ExcelWriter("RezultBot.xlsx")
    exl.write_to_sheet(run.main(), "ЖП")